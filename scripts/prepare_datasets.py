"""
Prepare public energy datasets into the project schema:

    time,meter_id,kwh,is_anomaly,source

The script is intentionally streaming-friendly for the London dataset. It writes a
sample by default so local experiments stay fast; pass --max-rows 0 to process all
rows found in the archive.
"""

from __future__ import annotations

import argparse
import csv
import zipfile
from datetime import datetime, timedelta
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"
PROCESSED_DIR = ROOT / "data" / "processed"


def normalize_header(name: str) -> str:
    return name.strip().lower().replace(" ", "_").replace(".", "")


def find_column(headers: list[str], candidates: tuple[str, ...]) -> str | None:
    normalized = {normalize_header(h): h for h in headers}
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]
    return None


def prepare_uci(max_rows: int) -> Path:
    archive = RAW_DIR / "individual_household_electric_power_consumption.zip"
    output = PROCESSED_DIR / "uci_meter_readings.csv"
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    if not archive.exists():
        raise FileNotFoundError(f"Missing {archive}. Run scripts/download_datasets.ps1 first.")

    written = 0
    with zipfile.ZipFile(archive) as zf, output.open("w", newline="", encoding="utf-8") as out:
        txt_name = next(name for name in zf.namelist() if name.endswith(".txt"))
        writer = csv.DictWriter(out, fieldnames=["time", "meter_id", "kwh", "is_anomaly", "source"])
        writer.writeheader()

        with zf.open(txt_name) as raw:
            reader = csv.DictReader((line.decode("latin-1") for line in raw), delimiter=";")
            for row in reader:
                active_power = row.get("Global_active_power")
                if not active_power or active_power == "?":
                    continue

                # Dataset value is kW at one-minute granularity, so kWh = kW / 60.
                kwh = float(active_power) / 60.0
                timestamp = datetime.strptime(f"{row['Date']} {row['Time']}", "%d/%m/%Y %H:%M:%S")
                writer.writerow(
                    {
                        "time": timestamp.isoformat(sep=" "),
                        "meter_id": "UCI_HOUSEHOLD_001",
                        "kwh": f"{kwh:.8f}",
                        "is_anomaly": "false",
                        "source": "uci_household_power",
                    }
                )
                written += 1
                if max_rows and written >= max_rows:
                    break

    print(f"[prepare] UCI rows written: {written} -> {output}")
    return output


def write_metadata(dataset_name: str, rows: int, meters: int, start_ts: str | None, end_ts: str | None, sampling_minutes: int | None, original_units: str | None, conversion: dict | None) -> None:
    import json
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    meta = {
        "dataset": dataset_name,
        "rows": rows,
        "distinct_meters": meters,
        "start_timestamp": start_ts,
        "end_timestamp": end_ts,
        "sampling_minutes": sampling_minutes,
        "original_units": original_units,
        "conversion": conversion,
    }
    with (PROCESSED_DIR / f"{dataset_name}_metadata.json").open("w", encoding="utf-8") as mf:
        json.dump(meta, mf, indent=2, ensure_ascii=False)


def prepare_nigeria(max_rows: int) -> Path:
    """Prepare the Nigerian household smart-meter dataset (HuggingFace mirror).

    This function loads the HF dataset (electricsheepafrica/nigerian_energy_and_utilities_household_smart_meter)
    and writes a processed CSV with the project schema using consumption_kwh as kwh.
    It also writes a metadata JSON and a short EDA markdown report.
    """
    try:
        from datasets import load_dataset
    except Exception:
        raise RuntimeError("Missing dependency: install 'datasets' to prepare Nigeria dataset.")

    output = PROCESSED_DIR / "nigeria_meter_readings.csv"
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    ds = load_dataset("electricsheepafrica/nigerian_energy_and_utilities_household_smart_meter")
    # single split 'train'
    ds_train = ds["train"]

    written = 0
    meters = set()
    min_ts = None
    max_ts = None

    with output.open("w", newline="", encoding="utf-8") as out:
        writer = csv.DictWriter(out, fieldnames=["time", "meter_id", "kwh", "is_anomaly", "source"])
        writer.writeheader()

        for i, row in enumerate(ds_train):
            if max_rows and max_rows > 0 and written >= max_rows:
                break
            ts = row.get("timestamp")
            meter = row.get("meter_id") or "NIG_UNKNOWN"
            kwh = row.get("consumption_kwh")
            if kwh is None:
                continue

            writer.writerow({
                "time": ts,
                "meter_id": meter,
                "kwh": f"{float(kwh):.8f}",
                "is_anomaly": "false",
                "source": "nigeria_smart_meter",
            })
            written += 1
            meters.add(meter)
            # update min/max timestamps (strings sortable)
            if ts:
                if min_ts is None or ts < min_ts:
                    min_ts = ts
                if max_ts is None or ts > max_ts:
                    max_ts = ts

    # infer sampling minutes: compute differences between first few timestamps per meter
    sampling = None
    try:
        from collections import defaultdict
        import datetime as _dt
        per_meter = defaultdict(list)
        for r in ds_train.select(range(min(2000, len(ds_train)))):
            per_meter[r["meter_id"]].append(r["timestamp"])
        # pick first meter with >=2 samples
        for m, ts_list in per_meter.items():
            if len(ts_list) >= 2:
                t0 = _dt.datetime.fromisoformat(ts_list[0])
                t1 = _dt.datetime.fromisoformat(ts_list[1])
                sampling = int((t1 - t0).total_seconds() / 60)
                break
    except Exception:
        sampling = None

    write_metadata("nigeria", written, len(meters), min_ts, max_ts, sampling, "kWh per interval (consumption_kwh)", None)

    # quick EDA write-up
    eda_path = Path(__file__).resolve().parents[1] / "reports" / "eda"
    eda_path.mkdir(parents=True, exist_ok=True)
    with (eda_path / "nigeria_eda.md").open("w", encoding="utf-8") as md:
        md.write(f"# Nigeria Smart Meter Dataset EDA\n\n")
        md.write(f"rows: {written}\\n\n")
        md.write(f"distinct_meters: {len(meters)}\\n\n")
        md.write(f"start_ts: {min_ts}\\n\n")
        md.write(f"end_ts: {max_ts}\\n\n")
        md.write(f"inferred_sampling_minutes: {sampling}\\n\n")

    print(f"[prepare] Nigeria rows written: {written} -> {output}")
    return output


def parse_london_start_timestamp(value: str) -> datetime:
    return datetime.strptime(value.strip(), "%Y-%m-%d %H-%M-%S")


def prepare_london_tsf(zf: zipfile.ZipFile, tsf_name: str, output: Path, max_rows: int) -> int:
    written = 0
    with output.open("w", newline="", encoding="utf-8") as out:
        writer = csv.DictWriter(out, fieldnames=["time", "meter_id", "kwh", "is_anomaly", "source"])
        writer.writeheader()

        with zf.open(tsf_name) as raw:
            in_data = False
            for raw_line in raw:
                line = raw_line.decode("utf-8", errors="replace").strip()
                if not line:
                    continue
                if line == "@data":
                    in_data = True
                    continue
                if not in_data or line.startswith("#") or line.startswith("@"):
                    continue

                try:
                    meter_id, start_value, values = line.split(":", 2)
                except ValueError:
                    continue

                start = parse_london_start_timestamp(start_value)
                for offset, value in enumerate(values.split(",")):
                    value = value.strip()
                    if not value or value == "?":
                        continue

                    timestamp = start + timedelta(minutes=30 * offset)
                    writer.writerow(
                        {
                            "time": timestamp.isoformat(sep=" "),
                            "meter_id": meter_id,
                            "kwh": value,
                            "is_anomaly": "false",
                            "source": "london_smart_meters",
                        }
                    )
                    written += 1
                    if max_rows and written >= max_rows:
                        return written

    return written


def prepare_london(max_rows: int) -> Path:
    archive = RAW_DIR / "london_smart_meters_dataset_without_missing_values.zip"
    output = PROCESSED_DIR / "london_meter_readings.csv"
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    if not archive.exists():
        raise FileNotFoundError(f"Missing {archive}. Run scripts/download_datasets.ps1 first.")

    written = 0
    with zipfile.ZipFile(archive) as zf:
        tsf_names = [name for name in zf.namelist() if name.lower().endswith(".tsf")]
        if tsf_names:
            written = prepare_london_tsf(zf, tsf_names[0], output, max_rows)
            print(f"[prepare] London rows written: {written} -> {output}")
            return output

        csv_names = [name for name in zf.namelist() if name.lower().endswith(".csv")]
        if not csv_names:
            raise RuntimeError(f"No CSV files found in {archive}")

    with zipfile.ZipFile(archive) as zf, output.open("w", newline="", encoding="utf-8") as out:
        writer = csv.DictWriter(out, fieldnames=["time", "meter_id", "kwh", "is_anomaly", "source"])
        writer.writeheader()

        for csv_name in csv_names:
            with zf.open(csv_name) as raw:
                text_iter = (line.decode("utf-8-sig", errors="replace") for line in raw)
                reader = csv.DictReader(text_iter)
                if not reader.fieldnames:
                    continue

                headers = reader.fieldnames
                meter_col = find_column(headers, ("lclid", "meter_id", "household_id", "id"))
                time_col = find_column(headers, ("tstp", "timestamp", "time", "datetime", "date_time"))
                kwh_col = find_column(headers, ("energy(kwh/halfhour)", "energy_kwh_halfhour", "kwh", "energy"))

                if not meter_col or not time_col or not kwh_col:
                    continue

                for row in reader:
                    value = (row.get(kwh_col) or "").strip()
                    if not value or value.lower() in {"null", "nan"}:
                        continue

                    writer.writerow(
                        {
                            "time": row[time_col],
                            "meter_id": row[meter_col],
                            "kwh": value,
                            "is_anomaly": "false",
                            "source": "london_smart_meters",
                        }
                    )
                    written += 1
                    if max_rows and written >= max_rows:
                        print(f"[prepare] London rows written: {written} -> {output}")
                        return output

    print(f"[prepare] London rows written: {written} -> {output}")
    return output


def prepare_morocco(max_rows: int) -> Path:
    """Prepare the UCI Morocco high-resolution smart-meter dataset.

    The UCI archive may contain an Excel file or CSVs. This function looks for
    a CSV first, otherwise attempts to find an .xlsx/.xls and raises a clear
    error if unsupported. Each row is normalized to the common schema.
    """
    archive = RAW_DIR / "morocco_high_resolution_smart_meters.zip"
    output = PROCESSED_DIR / "morocco_meter_readings.csv"
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    if not archive.exists():
        raise FileNotFoundError(f"Missing {archive}. Run scripts/download_datasets.ps1 first.")

    written = 0
    # The UCI archive may contain CSV or Excel files; we'll handle both below.
    with zipfile.ZipFile(archive) as zf:
        csv_names = [name for name in zf.namelist() if name.lower().endswith(".csv")]

    # The archive we have contains Excel files (xlsx). Use pandas/openpyxl if available
    try:
        import pandas as pd  # type: ignore
    except Exception as exc:  # pragma: no cover - environment-specific
        raise RuntimeError(
            "Preparing the Morocco dataset requires pandas and openpyxl. "
            "Please install them (pip install pandas openpyxl) and re-run."
        ) from exc

    # Prefer streaming parsing with openpyxl for memory and speed
    from openpyxl import load_workbook
    from io import BytesIO

    with zipfile.ZipFile(archive) as zf, output.open("w", newline="", encoding="utf-8") as out:
        writer = csv.DictWriter(out, fieldnames=["time", "meter_id", "kwh", "is_anomaly", "source"])
        writer.writeheader()

        xlsx_names = [name for name in zf.namelist() if name.lower().endswith(('.xlsx', '.xls'))]
        for xname in xlsx_names:
            with zf.open(xname) as raw:
                data = raw.read()
                bio = BytesIO(data)
                wb = load_workbook(filename=bio, read_only=True, data_only=True)
                # use first sheet
                ws = wb[wb.sheetnames[0]]
                rows = ws.iter_rows(values_only=True)
                try:
                    header = next(rows)
                except StopIteration:
                    continue

                # Normalize header names
                header_norm = [normalize_header(str(h)) for h in header]
                # Find time column index (usually first)
                time_idx = None
                for i, h in enumerate(header_norm):
                    if h in ("timestamp", "time", "datetime", "date", "date_time"):
                        time_idx = i
                        break
                if time_idx is None:
                    time_idx = 0

                value_indices = [i for i in range(len(header_norm)) if i != time_idx]
                if not value_indices:
                    continue

                # infer minutes from filename
                lower_name = xname.lower()
                if "30t" in lower_name or "30min" in lower_name:
                    minutes = 30
                elif "10t" in lower_name or "10min" in lower_name:
                    minutes = 10
                else:
                    # peek two rows to infer sampling interval
                    try:
                        first_row = next(rows)
                        second_row = next(rows)
                        import datetime as _dt
                        t1 = first_row[time_idx]
                        t2 = second_row[time_idx]
                        if isinstance(t1, _dt.datetime) and isinstance(t2, _dt.datetime):
                            minutes = int((t2 - t1).total_seconds() / 60)
                        else:
                            minutes = 10
                        # process the two rows below with a small buffer
                        buffer_rows = [first_row, second_row]
                    except StopIteration:
                        minutes = 10
                        buffer_rows = []

                period_hours = (minutes or 10) / 60.0

                # Build zone names from header
                zone_names = [header[i] if i < len(header) else f"zone{i}" for i in value_indices]

                # Keep original timestamps (10-min for most files, 30-min for Marrakech).
                # Convert amperes to estimated kW for non-Marrakech files using
                # Estimated_kW = 230 * I * 0.9 / 1000 = 0.207 * I
                is_marrakech = "marrakech" in xname.lower()

                # process any buffered rows first
                for br in locals().get("buffer_rows", []):
                    row_tuple = br
                    tcell = row_tuple[time_idx]
                    for col_i, zone in zip(value_indices, zone_names):
                        val = row_tuple[col_i]
                        if val is None:
                            continue
                        try:
                            v = float(val)
                        except Exception:
                            continue
                        if is_marrakech:
                            kW = v
                        else:
                            kW = 0.207 * v
                        kwh = kW * period_hours
                        try:
                            t_iso = tcell.isoformat(sep=" ") if hasattr(tcell, "isoformat") else str(tcell)
                        except Exception:
                            t_iso = str(tcell)
                        meter_id = f"{Path(xname).stem}_{normalize_header(str(zone))}"
                        writer.writerow({
                            "time": t_iso,
                            "meter_id": meter_id,
                            "kwh": f"{kwh:.8f}",
                            "is_anomaly": "false",
                            "source": "morocco_high_resolution",
                        })
                        written += 1

                # process remaining rows
                for row in rows:
                    row_tuple = row
                    tcell = row_tuple[time_idx]
                    for col_i, zone in zip(value_indices, zone_names):
                        val = row_tuple[col_i]
                        if val is None:
                            continue
                        try:
                            v = float(val)
                        except Exception:
                            continue
                        if is_marrakech:
                            kW = v
                        else:
                            kW = 0.207 * v
                        kwh = kW * period_hours
                        try:
                            t_iso = tcell.isoformat(sep=" ") if hasattr(tcell, "isoformat") else str(tcell)
                        except Exception:
                            t_iso = str(tcell)
                        meter_id = f"{Path(xname).stem}_{normalize_header(str(zone))}"
                        writer.writerow({
                            "time": t_iso,
                            "meter_id": meter_id,
                            "kwh": f"{kwh:.8f}",
                            "is_anomaly": "false",
                            "source": "morocco_high_resolution",
                        })
                        written += 1

    print(f"[prepare] Morocco rows written: {written} -> {output}")
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Prepare smart-grid datasets into a common CSV schema.")
    parser.add_argument("--dataset", choices=["london", "uci", "morocco", "nigeria", "all"], default="all")
    parser.add_argument("--max-rows", type=int, default=100_000, help="Rows per dataset; use 0 for all rows.")
    args = parser.parse_args()

    if args.dataset in {"london", "all"}:
        out = prepare_london(args.max_rows)
        # write simple metadata for London
        try:
            write_metadata("london", 0, 0, None, None, 30, "kWh per half-hour", None)
        except Exception:
            pass
    if args.dataset in {"uci", "all"}:
        out = prepare_uci(args.max_rows)
        try:
            write_metadata("uci", 0, 1, None, None, 1, "kWh per minute (from kW)", None)
        except Exception:
            pass
    if args.dataset in {"morocco", "all"}:
        out = prepare_morocco(args.max_rows)
        try:
            # Morocco has mixed sampling; leave sampling_minutes null in metadata
            write_metadata("morocco", 0, 0, None, None, None, "mixed: some files in amperes (converted), Marrakech in kW", {"ampere_to_kW_factor": 0.207})
        except Exception:
            pass
    if args.dataset in {"nigeria", "all"}:
        out = prepare_nigeria(args.max_rows)


if __name__ == "__main__":
    main()
